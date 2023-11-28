import requests
import openpyxl

def get_warranty_info(sn):
    cookies = {
    'sce': '1',
    'Hm_lvt_062eaba7adfd61ea654636b61c45cf36': '1701063372',
    'leid': '1.8acTrQfO2iY',
    'Hm_lpvt_062eaba7adfd61ea654636b61c45cf36': '1701067778',
    }
    headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36'}

    payload = {'search_key': sn}
    r = requests.get('https://newthink.lenovo.com.cn/api/ThinkHome/Machine/DriveMachineListInfo', params=payload, cookies=cookies, headers=headers,).json()
    rstatusCode = r["statusCode"]
    if rstatusCode != 200:
        return (sn,"联想官网未查询到记录,请检查输入sn是否正确")

    params = {'sn': sn}
    response = requests.get('https://newthink.lenovo.com.cn/api/ThinkHome/Machine/WarrantyListInfo', params=params, cookies=cookies, headers=headers,).json()

    # 保修开始时间和结束时间，如果为空则取onsitestart
    start_time = response["data"]["base_data"][0]["laborstart_date"]
    end_time = response["data"]["base_data"][0]["laborend_date"]
    if start_time is None:
        start_time = response["data"]["base_data"][0]["onsitestart_date"]
    if end_time is None:
        end_time = response["data"]["base_data"][0]["onsiteend_date"]

    # 剩余保修天数
    warranty_day = response["data"]["base_data"][0]["warranty_day"]
    if warranty_day >0:
        warranty_day = ("未过保剩余", warranty_day)
    else:
        warranty_day = ("已过保", warranty_day)

    return (start_time, end_time,warranty_day)

def warranty_2_excel(path):
    error = {}
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws['N1'] = "保修开始时间"
    ws['O1'] = "保修结束时间"
    ws['P1'] = "保修状态"
    for i in range(2, ws.max_row+1):
        sn = str(ws['C'+str(i)].value)  # 获取sn字段
        print(f"正在查询 {sn}")
        if sn is None:
            continue
        if '\u4e00' <= sn <= '\u9fff':
            print("输入字符为中文")
            continue
        try:
            value = get_warranty_info(sn)
        except TypeError as e:
            error[sn] = e
        ws['N' + str(i)].value = value[0]
        ws['O' + str(i)].value = value[1]
        if len(value) == 3:
            ws['P' + str(i)].value = str(value[2])
    wb.save(path)
    if len(error) > 0:
        print(error)
if __name__ == '__main__':
    path = '/Users/xxx/lenovo.xlsx'
    warranty_2_excel(path) 
